# coding: utf-8
import openpyxl
from openpyxl import load_workbook
import os ,time
from urllib.parse import quote_plus as urlquote
from sqlalchemy import create_engine
import pandas as pd
import json
from flask import Flask, request
import datetime

app = Flask(__name__)
ex_path = 'testlog.xlsx'

def new_excel():
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    for i in list1:
        sheet.cell(i[0], i[1]).value = i[2]
    workbook.save(ex_path)
    time.sleep(0.01)
    excel_to_mysql()

def excel_to_mysql():
    userName = "root"
    password = "zz@123456"
    dbHost = "127.0.0.1"
    dbPort = 3306
    dbName = "unicom"
    engine = create_engine(f'mysql+pymysql://{userName}:{urlquote(password)}@{dbHost}:{dbPort}/{dbName}?charset=utf8')
    data_frame = pd.read_excel(ex_path)
    data_frame = data_frame.where(data_frame.notnull(), '')
    data_frame.to_sql(name='testlog', con=engine, index=False, if_exists='replace')

@app.route('/test',methods=['GET','POST'])
def index():
    global list1
    if request.method == "POST":
        a = request.form.get('exportdata')
        b = json.loads(a)
        for key, data in b.items():
            if key == 'celldata':
                list1 = []
                for i in data:
                    for j in i['v']:
                        if j == 'm':
                            row = i['r'] + 2
                            col = i['c'] + 1
                            value = i['v']['m']
                            list1.extend([[row, col, value]])
                if os.path.exists(ex_path) == False:
                    new_excel()
                if os.path.exists(ex_path) == True:
                    os.remove(ex_path)
                    new_excel()

if __name__ == '__main__':
    app.run(host="127.0.0.1", port=5000)
    #app.run(host="0.0.0.0", port=5000)